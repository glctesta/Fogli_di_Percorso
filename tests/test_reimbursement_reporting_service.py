"""Test service per report rimborsi effettivi."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from fdp_app.reimbursement_reporting.service import (
    build_reimbursement_xlsx,
    effective_reimbursement,
)
from fdp_app.repos.reimbursement_reporting_repo import ReimbursementReportRow


def _row(**kwargs) -> ReimbursementReportRow:
    defaults = dict(
        employee_hire_history_id=101,
        employee_surname="Rossi",
        employee_name="Mario",
        declared_amount_eur=120.0,
        additional_amount_eur=15.0,
        deduction_amount_eur=5.0,
        notes="Aggiunta manuale",
        last_updated_on=datetime(2026, 6, 1, 10, 30),
    )
    defaults.update(kwargs)
    return ReimbursementReportRow(**defaults)


def test_effective_reimbursement_formula():
    row = _row(declared_amount_eur=100.0, additional_amount_eur=20.0, deduction_amount_eur=30.0)
    assert effective_reimbursement(row) == 90.0


def test_build_reimbursement_xlsx_contains_effective_amount():
    xlsx = build_reimbursement_xlsx([_row()], year=2026, month=5, sheet_name="Rimborsi")
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active

    assert ws.cell(row=2, column=2).value == "Rossi"
    assert ws.cell(row=2, column=7).value == 130.0
    assert ws.cell(row=4, column=1).value == "TOTALE"
    assert ws.cell(row=4, column=7).value == 130.0


def test_build_reimbursement_xlsx_handles_empty_rows():
    xlsx = build_reimbursement_xlsx([], year=2026, month=5, sheet_name="Rimborsi")
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active

    # Header + totals + metadata rows, no data rows.
    assert ws.cell(row=1, column=1).value == "EmployeeHireHistoryId"
    assert ws.cell(row=3, column=1).value == "TOTALE"
    assert ws.cell(row=3, column=7).value == 0.0
    assert ws.cell(row=5, column=1).value == "Periodo: 05/2026"
