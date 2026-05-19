"""Test del modulo admin: generazione XLSX + endpoint /admin/export."""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from fdp_app.admin.service import build_xlsx
from fdp_app.repos.pathtrack_repo import PathTrackRow, PathTrackWithEmployee


def _entry(employee_surname="Rossi", employee_name="Mario", **kwargs):
    defaults = dict(
        path_track_id=1, registry_id=500, date_path_track=date(2026, 4, 1),
        declarated_path_id=99, in_behalf_of_id=None,
        reimbursement_type="CARBURANTE", number_of_trips=20, road_km=10.0,
        rate_id_used=3, taxi_total_eur=None, computed_amount_eur=53.55,
        status="SUBMITTED", submitted_on=datetime(2026, 5, 3, 10, 0),
    )
    defaults.update(kwargs)
    return PathTrackWithEmployee(
        row=PathTrackRow(**defaults),
        employee_surname=employee_surname,
        employee_name=employee_name,
    )


def test_build_xlsx_generates_valid_workbook():
    entries = [
        _entry(),
        _entry(employee_surname="Bianchi", employee_name="Luigi"),
    ]
    xlsx_bytes = build_xlsx(entries, year=2026, month=4, month_name="Aprile")
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100  # not empty

    # Open it back to verify content
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "Aprile 2026"
    # Header
    header_row = [c.value for c in ws[1]]
    assert "Cognome" in header_row
    assert "Importo EUR" in header_row
    # Data rows
    assert ws.cell(row=2, column=1).value == "Rossi"
    assert ws.cell(row=3, column=1).value == "Bianchi"


def test_build_xlsx_handles_empty_rows():
    xlsx_bytes = build_xlsx([], year=2026, month=4, month_name="Aprile")
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    # Only header row
    assert ws.max_row == 1


def test_build_xlsx_handles_draft_with_null_registry_and_submitted_on():
    entry = _entry(status="DRAFT", registry_id=None, submitted_on=None)
    xlsx_bytes = build_xlsx([entry], year=2026, month=5, month_name="Maggio")
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    # Registry column should be empty string
    assert ws.cell(row=2, column=8).value in (None, "")
    assert ws.cell(row=2, column=9).value in (None, "")


# ---- /admin/export endpoint ----

@pytest.fixture
def mock_pt_repo_export():
    with patch("fdp_app.admin.routes.PathTrackRepo") as cls:
        instance = MagicMock()
        instance.list_for_sub_cdc.return_value = []
        cls.return_value = instance
        yield instance


def _login_admin(client, eh_id=10, sub_cdc_id=42, fc=70):
    with client.session_transaction() as sess:
        sess["user_id"] = eh_id
        sess["full_name"] = "Rossi Mario"
        sess["sub_cdc_id"] = sub_cdc_id
        sess["function_code"] = fc


def test_export_requires_admin(client):
    with client.session_transaction() as sess:
        sess["user_id"] = 10
        sess["function_code"] = 50
    response = client.get("/admin/export?year=2026&month=4")
    assert response.status_code == 403


def test_export_returns_xlsx_with_correct_filename(client, mock_pt_repo_export):
    _login_admin(client)
    response = client.get("/admin/export?year=2026&month=4")
    assert response.status_code == 200
    assert response.mimetype.startswith("application/vnd.openxmlformats")
    cd = response.headers.get("Content-Disposition", "")
    assert "fogli-di-percorso-2026-04.xlsx" in cd


def test_export_rejects_invalid_month(client, mock_pt_repo_export):
    _login_admin(client)
    response = client.get("/admin/export?year=2026&month=13")
    assert response.status_code == 400


def test_export_rejects_missing_year(client, mock_pt_repo_export):
    _login_admin(client)
    response = client.get("/admin/export?month=4")
    assert response.status_code == 400
