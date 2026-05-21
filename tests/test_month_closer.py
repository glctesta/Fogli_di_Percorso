"""Test del MonthCloser (genera XLSX dei dipendenti pendenti)."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest
from flask import Flask
from openpyxl import load_workbook

from fdp_app.notifications.service import MonthCloser
from fdp_app.pathtracks.deadline import previous_month_first_day
from fdp_app.repos.employee_repo import PendingEmployee


def _make_flask_app(tmp_path: Path) -> Flask:
    root = tmp_path / "fdp_app"
    root.mkdir(parents=True, exist_ok=True)
    app = Flask(__name__, root_path=str(root))
    app.config["_db"] = MagicMock()

    class FakeSettings:
        APP_URL = "http://test.local"

    app.config["_settings_cls"] = FakeSettings
    return app


def _patch_repo_pending(monkeypatch, pending_list):
    import fdp_app.notifications.service as svc_mod

    fake_repo_cls = MagicMock()
    fake_repo = MagicMock()
    fake_repo.find_pending_for_month.return_value = pending_list
    fake_repo_cls.return_value = fake_repo
    monkeypatch.setattr(svc_mod, "EmployeeRepo", fake_repo_cls)


def test_run_with_empty_pending_creates_header_only_xlsx(tmp_path, monkeypatch):
    app = _make_flask_app(tmp_path)
    _patch_repo_pending(monkeypatch, [])
    closer = MonthCloser(app)

    out_path, count = closer.run()

    assert isinstance(out_path, Path)
    assert out_path.exists()
    assert count == 0
    wb = load_workbook(out_path)
    ws = wb.active
    # max_row is 1 because we only appended the header.
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == ["Cognome", "Nome", "Email"]


def test_run_with_three_pending_creates_three_data_rows(tmp_path, monkeypatch):
    app = _make_flask_app(tmp_path)
    pending = [
        PendingEmployee(101, "Bianchi", "Luigi", "lbianchi@example.com"),
        PendingEmployee(102, "Verdi", "Maria", "mverdi@example.com"),
        PendingEmployee(103, "Rossi", "Mario", "mrossi@example.com"),
    ]
    _patch_repo_pending(monkeypatch, pending)
    closer = MonthCloser(app)

    out_path, count = closer.run()

    assert count == 3
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.max_row == 4  # 1 header + 3 rows
    row2 = [c.value for c in ws[2]]
    assert row2 == ["Bianchi", "Luigi", "lbianchi@example.com"]
    row4 = [c.value for c in ws[4]]
    assert row4 == ["Rossi", "Mario", "mrossi@example.com"]


def test_run_header_styled_with_navy_fill_and_white_bold_font(tmp_path, monkeypatch):
    app = _make_flask_app(tmp_path)
    _patch_repo_pending(monkeypatch, [])
    closer = MonthCloser(app)

    out_path, _ = closer.run()

    wb = load_workbook(out_path)
    ws = wb.active
    for cell in ws[1]:
        # openpyxl stores ARGB internally; check rgb attribute contains hex.
        fill_rgb = cell.fill.fgColor.rgb or ""
        # 0B2A5B navy -> stored as '000B2A5B' (ARGB) in many openpyxl versions.
        assert "0B2A5B" in str(fill_rgb).upper()
        assert cell.font.bold is True
        font_rgb = cell.font.color.rgb if cell.font.color else ""
        assert "FFFFFF" in str(font_rgb).upper()


def test_run_worksheet_title_contains_target_month_yyyy_mm(tmp_path, monkeypatch):
    app = _make_flask_app(tmp_path)
    _patch_repo_pending(monkeypatch, [])
    closer = MonthCloser(app)

    out_path, _ = closer.run()

    target = previous_month_first_day()
    expected_yyyy_mm = f"{target:%Y-%m}"
    wb = load_workbook(out_path)
    ws = wb.active
    assert expected_yyyy_mm in ws.title


def test_run_sets_column_widths(tmp_path, monkeypatch):
    app = _make_flask_app(tmp_path)
    pending = [PendingEmployee(1, "Bianchi", "Luigi", "lbianchi@example.com")]
    _patch_repo_pending(monkeypatch, pending)
    closer = MonthCloser(app)

    out_path, _ = closer.run()

    wb = load_workbook(out_path)
    ws = wb.active
    # Each of the 3 columns should have a width attribute > 0.
    for letter in ("A", "B", "C"):
        dim = ws.column_dimensions[letter]
        assert dim.width is not None
        assert dim.width > 0
