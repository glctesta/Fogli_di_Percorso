"""Servizi notifiche: reminder email, close-month report, BNR refresh."""
from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path
from typing import Tuple

from dateutil.relativedelta import relativedelta
from jinja2 import Environment, FileSystemLoader

from fdp_app.pathtracks.deadline import previous_month_first_day
from fdp_app.repos.employee_repo import EmployeeRepo
from email_connector import EmailSender

_logger = logging.getLogger("fdp.notifications")

_MONTH_NAMES = {
    "ro": ["", "Ianuarie", "Februarie", "Martie", "Aprilie", "Mai", "Iunie",
           "Iulie", "August", "Septembrie", "Octombrie", "Noiembrie", "Decembrie"],
    "it": ["", "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
           "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"],
    "en": ["", "January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"],
}

_VALID_STAGES = ("opening", "midway", "last-call")


class EmailReminderService:
    """Orchestra l'invio di reminder email per un determinato stage.

    Idempotenza: traccia il run giornaliero tramite un file
    `state/reminders-{YYYY-MM-DD}-{stage}.done` contenente il riassunto JSON.
    """

    def __init__(self, app, default_lang: str = "ro") -> None:
        self._app = app
        self._default_lang = default_lang
        self._state_dir = Path(app.root_path).parent / "state"
        self._state_dir.mkdir(exist_ok=True)
        self._jinja = Environment(
            loader=FileSystemLoader(
                str(Path(app.root_path) / "notifications" / "templates" / "email")
            ),
            autoescape=False,  # plain-text emails
        )

    def _state_file(self, stage: str) -> Path:
        today = date.today().isoformat()
        return self._state_dir / f"reminders-{today}-{stage}.done"

    def _already_sent_today(self, stage: str) -> bool:
        return self._state_file(stage).exists()

    def _mark_sent(self, stage: str, summary: dict) -> None:
        self._state_file(stage).write_text(
            json.dumps(summary, indent=2), encoding="utf-8"
        )

    def run_stage(self, stage: str) -> Tuple[int, int]:
        """Esegue lo stage. Ritorna `(sent, skipped)`.

        Idempotente: se lo stage e' gia' stato eseguito oggi (file `.done`
        presente), ritorna `(0, 0)` senza inviare nulla.
        """
        if stage not in _VALID_STAGES:
            raise ValueError(f"stage non valido: {stage}")

        if self._already_sent_today(stage):
            _logger.info("Stage %s gia' inviato oggi, skip", stage)
            return 0, 0

        target_month = previous_month_first_day()
        target_year = target_month.year
        next_month = target_month + relativedelta(months=1)

        db = self._app.config["_db"]
        repo = EmployeeRepo(db)
        pending = repo.find_pending_for_month(date_path_track=target_month)

        if not pending:
            _logger.info("Nessun dipendente pendente per %s", target_month)
            self._mark_sent(stage, {"sent": 0, "skipped": 0, "pending": 0})
            return 0, 0

        lang = self._default_lang
        template_name = f"reminder_{stage.replace('-', '_')}_{lang}.txt"
        try:
            template = self._jinja.get_template(template_name)
        except Exception as e:
            _logger.error("Template %s non trovato: %s", template_name, e)
            raise

        sender = EmailSender()
        sent_count = 0
        skipped_count = 0
        results = []

        app_url = self._app.config["_settings_cls"].APP_URL

        for emp in pending:
            try:
                rendered = template.render(
                    full_name=emp.full_name,
                    month_name=_MONTH_NAMES[lang][target_month.month],
                    year=target_year,
                    next_month_name=_MONTH_NAMES[lang][next_month.month],
                    next_year=next_month.year,
                    app_url=app_url,
                )
                lines = rendered.strip().split("\n")
                subject = lines[0].replace("Subject:", "").strip()
                body = "\n".join(lines[2:]).strip()
                sender.send_email(emp.work_email, subject, body, is_html=False)
                sent_count += 1
                results.append({
                    "employee_id": emp.employee_hire_history_id,
                    "email": emp.work_email,
                    "status": "sent",
                })
                _logger.info("Reminder %s inviato a %s (%s)",
                             stage, emp.full_name, emp.work_email)
            except Exception as e:
                skipped_count += 1
                results.append({
                    "employee_id": emp.employee_hire_history_id,
                    "email": emp.work_email,
                    "status": "error",
                    "error": str(e),
                })
                _logger.error("Reminder %s fallito per %s: %s",
                              stage, emp.full_name, e)

        self._mark_sent(stage, {
            "sent": sent_count,
            "skipped": skipped_count,
            "pending": len(pending),
            "results": results,
        })
        return sent_count, skipped_count


class MonthCloser:
    """Genera un XLSX con i dipendenti che non hanno inviato la dichiarazione
    per il mese precedente."""

    def __init__(self, app) -> None:
        self._app = app
        self._state_dir = Path(app.root_path).parent / "state"
        self._state_dir.mkdir(exist_ok=True)

    def run(self) -> Tuple[Path, int]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        target_month = previous_month_first_day()
        db = self._app.config["_db"]
        repo = EmployeeRepo(db)
        pending = repo.find_pending_for_month(date_path_track=target_month)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Mancanti {target_month:%Y-%m}"
        ws.append(["Cognome", "Nome", "Email"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B2A5B")
        for emp in pending:
            ws.append([emp.surname, emp.name, emp.work_email])
        for col in ws.columns:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(length + 2, 50)

        out = self._state_dir / f"missing-{target_month:%Y-%m}.xlsx"
        wb.save(out)
        _logger.info("close-month: %s pendenti, file=%s", len(pending), out)
        return out, len(pending)


class BnrRefreshJob:
    """Pre-popola il tasso BNR del giorno corrente nella tabella cache."""

    def __init__(self, app) -> None:
        self._app = app

    def run(self) -> Tuple[float, str]:
        from fdp_app.pathtracks.currency import CurrencyService
        from fdp_app.pathtracks.bnr_client import BnrRateClient
        from fdp_app.repos.bnr_rate_repo import BnrRateRepo

        db = self._app.config["_db"]
        client = self._app.config.get("_bnr_client") or BnrRateClient()
        service = CurrencyService(bnr_repo=BnrRateRepo(db), bnr_client=client)
        resolved = service.resolve_for(date.today(), user_sys="bnr-refresh-job")
        _logger.info("BNR refresh: %s = %s",
                     resolved.source, resolved.value_ron_per_eur)
        return resolved.value_ron_per_eur, resolved.source
