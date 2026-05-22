"""Helpers admin: generazione XLSX in memoria."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_xlsx(
    rows,
    *,
    year: int,
    month: int,
    month_name: str,
    truncated: bool = False,
    max_rows: int | None = None,
) -> bytes:
    """Genera un XLSX con una riga per PathTrackWithEmployee.

    rows: iterable di PathTrackWithEmployee
    truncated: True se la lista è stata troncata al cap. Inserisce una riga
        banner all'inizio del foglio.
    max_rows: cap utilizzato (richiesto quando truncated=True, ignorato altrimenti)
    """
    if truncated and max_rows is None:
        raise ValueError("max_rows must be set when truncated=True")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    headers = [
        "Cognome", "Nome", "Tipo", "Stato", "N. Viaggi A/R", "Km one-way",
        "Importo EUR", "N. Registro", "Data invio",
    ]

    if truncated:
        banner_text = (
            f"AVVISO: risultati troncati a {max_rows} righe. "
            "Restringere i filtri per dati completi."
        )
        ws.append([banner_text])
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(headers),
        )
        banner_cell = ws.cell(row=1, column=1)
        banner_cell.font = Font(bold=True, color="000000")
        banner_cell.fill = PatternFill("solid", fgColor="FFF3CD")
        banner_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_row_idx = 2
    else:
        header_row_idx = 1

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B2A5B")
    header_align = Alignment(horizontal="center")
    for cell in ws[header_row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for entry in rows:
        ws.append([
            entry.employee_surname,
            entry.employee_name,
            entry.row.reimbursement_type,
            entry.row.status,
            entry.row.number_of_trips,
            entry.row.road_km,
            entry.row.computed_amount_eur,
            entry.row.registry_id or "",
            entry.row.submitted_on.strftime("%d/%m/%Y %H:%M") if entry.row.submitted_on else "",
        ])

    # Auto-width approssimativo. Skip the banner row (merged cells confuse the
    # per-column max-length calc), so iterate from the header row down.
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row_idx in range(header_row_idx, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
