"""Service helpers per report rimborsi ed export Excel."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fdp_app.repos.reimbursement_reporting_repo import ReimbursementReportRow


REPORT_HEADERS = [
    "EmployeeHireHistoryId",
    "Cognome",
    "Nome",
    "Importo dichiarato EUR",
    "Integrazione EUR",
    "Detrazione EUR",
    "Rimborso effettivo EUR",
    "Note",
    "Ultimo aggiornamento",
]


def effective_reimbursement(row: ReimbursementReportRow) -> float:
    return row.declared_amount_eur + row.additional_amount_eur - row.deduction_amount_eur


def build_reimbursement_xlsx(
    rows: list[ReimbursementReportRow],
    *,
    year: int,
    month: int,
    sheet_name: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(REPORT_HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B2A5B")
    header_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    total_declared = 0.0
    total_additional = 0.0
    total_deduction = 0.0
    total_effective = 0.0

    for row in rows:
        effective = effective_reimbursement(row)
        total_declared += row.declared_amount_eur
        total_additional += row.additional_amount_eur
        total_deduction += row.deduction_amount_eur
        total_effective += effective
        ws.append([
            row.employee_hire_history_id,
            row.employee_surname,
            row.employee_name,
            row.declared_amount_eur,
            row.additional_amount_eur,
            row.deduction_amount_eur,
            effective,
            row.notes,
            row.last_updated_on.strftime("%d/%m/%Y %H:%M") if row.last_updated_on else "",
        ])

    total_row_idx = ws.max_row + 2
    ws.cell(row=total_row_idx, column=1).value = "TOTALE"
    ws.cell(row=total_row_idx, column=4).value = round(total_declared, 2)
    ws.cell(row=total_row_idx, column=5).value = round(total_additional, 2)
    ws.cell(row=total_row_idx, column=6).value = round(total_deduction, 2)
    ws.cell(row=total_row_idx, column=7).value = round(total_effective, 2)

    for col_idx in (1, 4, 5, 6, 7):
        ws.cell(row=total_row_idx, column=col_idx).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=col_idx).fill = PatternFill("solid", fgColor="E2E3E5")

    for col_idx in range(1, len(REPORT_HEADERS) + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 55)

    ws.freeze_panes = "A2"

    # Metadata minimale utile per audit export.
    ws.cell(row=ws.max_row + 2, column=1).value = f"Periodo: {month:02d}/{year}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
